import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
import os
import webbrowser
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from datetime import datetime

class IDR_1(tk.Tk):
    def __init__(self, window_size="800x600"):
        super().__init__()
        self.title("Daily Inbound Report Generator")
        self.geometry(window_size)


        # Create a textbox with a scrollbar for instructions
        self.instructions_text = tk.Text(self, wrap="word", bg="white", height=10)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.instructions_text.yview)
        self.instructions_text.configure(yscrollcommand=self.scrollbar.set)

        # Insert guidance text
        instructions = r"""Retrieve the following WMS data and import to generate daily inbound report:
[Weekly Inbound Daily Plan] - copy from [N:\8301\Store\Daily Plan\Inbound]
[Receiving] - [Activity Tracking] - Transaction Type = Receiving Fuction
[Rework] - [Activity Tracking] - Transaction Type = Inventory Movement, Transaction Code = Locate Case (Non-Directed)
[Putaway] - [Activity Tracking] - Transaction Type = Pulling/Replenishment Functions, Transaction Code = Replenished Active
[Adhoc] - [Activity Tracking] - Transaction Type = Pulling/Replenishment Functions, Transaction Code = Pack Case from Active
[Transfer / Debit] - [Activity Tracking] - Transaction Type = Packing, Transaction Code = Packing (dir), Menu Option Name = DJ Pick Dbt/Tfr/Mdo
[Cycel Count] - [Activity Tracking] - Transaction Type = Miscellaneous, Transaction Code = Cycle Count: Locations counted.
[ASN] - [Inbound Shipment Inquiry] - set [ASN Verified Date]
[Received units by dept.] - [Inbound Shipment Inquiry] - set [Rcvd Date From] and [Rcvd Date To]""" + "\n\n"

        self.instructions_text.insert("1.0", instructions)
        self.instructions_text.config(state='disabled')

        # Layout the textbox and scrollbar
        self.instructions_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=0, column=1, sticky="ns", pady=10)

        # Configure grid row/column weights to ensure the textbox expands with the window size
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.receving_label = tk.Label(self, text="Acceptable delay for ASN (minutes):")
        self.receving_label.grid(row=1, column=0, padx=10, pady=10, sticky="w")
        self.receving_label = tk.Label(self, text="Acceptable delay for Receiving (minutes):")
        self.receving_label.grid(row=2, column=0, padx=10, pady=10, sticky="w")
        self.receving_label = tk.Label(self, text="Acceptable delay for Putaway/Adhoc (minutes):")
        self.receving_label.grid(row=3, column=0, padx=10, pady=10, sticky="w")
        self.receving_label = tk.Label(self, text="Acceptable delay for Rework (units/minutes):")
        self.receving_label.grid(row=4, column=0, padx=10, pady=10, sticky="w")
        self.receving_label = tk.Label(self, text="Acceptable delay for Transfer/Debit (minutes):")
        self.receving_label.grid(row=5, column=0, padx=10, pady=10, sticky="w")
        self.receving_label = tk.Label(self, text="Acceptable delay for Cycle Count (minutes):")
        self.receving_label.grid(row=6, column=0, padx=10, pady=10, sticky="w")

        self.time_delays = []
        for i in range(6):
            entry = ttk.Entry(self)
            entry.grid(row=1+i, column=0, padx=10, pady=10, sticky="ns")
            if i == 3: entry.insert(0, "60")
            else: entry.insert(0, "15")
            self.time_delays.append(entry)     

        # Create the "Import Excel Files" button and layout using grid
        self.import_button = ttk.Button(self, text="Import Excel Files", command=lambda: self.import_excel(self.time_delays))
        self.import_button.grid(row=7, column=0, padx=10, pady=10, sticky="e")  

    def split_and_calculate_differences(self,times, acceptable_delay):
        results = []
        differences = []  # 存储每个子列表的首尾差值
        current_list = [times[0]]  # 开始新的子列表
        acceptable_delay_int = int(acceptable_delay)

        for i in range(1, len(times)):
            if times[i] - times[i - 1] > acceptable_delay_int:
                # 结束当前子列表，计算差值
                results.append(current_list)
                differences.append(((current_list[-1] - current_list[0])//15+1)/4)
                current_list = [times[i]]  # 开始新的子列表
            else:
                current_list.append(times[i])  # 继续填充当前子列表

        results.append(current_list)
        differences.append(((current_list[-1] - current_list[0])//15+1)/4)

        # print(results)
        # print(differences)

        return differences

    def calculate_user_times(self,df_activity,acceptable_delay):
        df_activity['Date Time'] = pd.to_datetime(df_activity['Date Time'], format='%d/%m/%Y %H:%M')
        df_activity['Time Value'] = df_activity['Date Time'].dt.hour * 60 + df_activity['Date Time'].dt.minute
        user_times = {}
        for user, group in df_activity.groupby('User'):
            user_times[user] = sorted(group['Time Value'].tolist())

        # print(user_times)

        for user, times in user_times.items():
            sum_differences = self.split_and_calculate_differences(times, acceptable_delay)
            user_times[user] = sum_differences
        # print(user_times)
        user_times = {user: sum(values) for user, values in user_times.items()}
        return user_times
    
     
    def import_excel(self,delays):
        self.instructions_text.config(state='normal')
        interval_acceptable_delay = [entry.get() for entry in delays]
        # print("Entry values:", interval_acceptable_delay)
        self.instructions_text.insert(tk.END, "\nStarting to import data...\n")
        file_paths = list(filedialog.askopenfilenames(filetypes=[("Excel files", "*.xls;*.xlsx")]))
        # print(file_paths)

        file_data = { # 预分类文件
            'Inbound Daily Plan': [],
            'Receiving': [],
            'Rework': [],
            'Transfer / Debit': [],
            'Putaway': [],
            'Adhoc': [],
            'Cycle Count': [],
            'ASN': [],
            'ASN verified': [],
            'Received by Group': [],
            'Other': []
        }

        for file_path in file_paths:
            # print (file_path)
            df = pd.read_excel(file_path, nrows=2)  # 只加载前两行进行检查
            if "Inbound Daily Plan" in file_path:
                file_data['Inbound Daily Plan'].append(file_path)
            elif 'Receiving Functions' in df.iloc[1].to_string():
                file_data['Receiving'].append(file_path)
            elif 'Inventory Movement' in df.iloc[1].to_string() and 'Locate Case (Non-Directed)' in df.iloc[1].to_string():
                file_data['Rework'].append(file_path)
            elif 'DJ Pick Dbt/Tfr/Mdo' in df.iloc[1].to_string() and 'Packing (dir)' in df.iloc[1].to_string():
                file_data['Transfer / Debit'].append(file_path)
            elif 'Replenished Active' in df.iloc[1].to_string():
                file_data['Putaway'].append(file_path)
            elif 'Pack Case from Active' in df.iloc[1].to_string():
                file_data['Adhoc'].append(file_path)
            elif 'Cycle Count: Locations counted.' in df.iloc[1].to_string():
                file_data['Cycle Count'].append(file_path)
            else: file_data['ASN'].append(file_path)

        for asn_path in file_data['ASN']:
            df = pd.read_excel(asn_path, usecols=['Status'])
            if (df.iloc[:, 0] == 'Receipt Verified').all():
                file_data['ASN verified'].append(asn_path)
            elif df.iloc[:, 0].isin(['In Receiving', 'Receipt Verified']).all():
                file_data['Received by Group'].append(asn_path)
            else: file_data['Other'].append(asn_path)

        for category, files_list in file_data.items():
            if category not in ['ASN', 'Other'] and len(files_list) > 0:
                self.instructions_text.insert('end', f'\nFiles list for {category}:\n')
                for item in files_list:
                    self.instructions_text.insert('end', item + '\n')

        self.instructions_text.see('end')
        self.instructions_text.config(state='disabled')

        df_receiving_result, df_pa_result, df_ah_result, df_rework_result, df_trf_result, df_dbt_result, df_cc_result, df_asn_result = ({} for _ in range(8))
        total_receiving_units, total_receiving_cases, total_receiving_units_food, total_rework_units, \
        total_pa_units, total_pa_units_food, total_ah_units, total_ah_units_food, \
        total_trf_units, total_dbt_units, total_trf_units_food, total_dbt_units_food, total_asn = (0 for _ in range(13))
        day_of_week = None

        if len(file_data['Receiving']) > 0:
            df_receiving = pd.read_excel(file_data['Receiving'][0], usecols=["Date Time", "User", "Container","Dept","Nbr Units"],dtype=str)
            day_of_week = pd.to_datetime(df_receiving.loc[0, 'Date Time'], format='%d/%m/%Y %H:%M').day_name()
            
            df_receiving.columns = df_receiving.columns.str.strip()
            df_receiving_food_sure = df_receiving[df_receiving['Dept'].astype(str).str.startswith('0')]
            df_receiving_food_maybe = df_receiving[(df_receiving['Dept'].isna()) & (df_receiving['Container'].str.startswith('1000')) & (df_receiving['Container'].str.len() == 13)]
            df_receiving = df_receiving.drop(df_receiving_food_sure.index)
            df_receiving_food_maybe = df_receiving_food_maybe.reset_index(drop=True)
            df_receiving_result = self.calculate_user_times(df_receiving,interval_acceptable_delay[1])
            df_receiving['Nbr Units'] = pd.to_numeric(df_receiving['Nbr Units'], errors='coerce').fillna(0).astype(int)
            df_receiving_food_sure['Nbr Units'] = pd.to_numeric(df_receiving_food_sure['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_receiving_units = df_receiving['Nbr Units'].sum()
            total_receiving_cases = df_receiving['Container'].nunique()
            total_receiving_units_food = df_receiving_food_sure['Nbr Units'].sum()

        if len(file_data['Putaway']) > 0:
            df_pa = pd.concat([pd.read_excel(file, usecols=["Date Time", "User", "Container", "Dept", "Nbr Units", "To Location"], dtype=str) for file in file_data['Putaway']], ignore_index=True)
            day_of_week = pd.to_datetime(df_pa.loc[0, 'Date Time'], format='%d/%m/%Y %H:%M').day_name()
            
            df_pa.columns = df_pa.columns.str.strip()
            df_pa_food = df_pa[df_pa['Dept'].astype(str).str.startswith('0')]
            df_pa = df_pa.drop(df_pa_food.index)

            df_pa['Location'] = df_pa['To Location'].astype(str).str[:2]
            df_pa['Nbr Units'] = df_pa['Nbr Units'].astype(int)
            pivot_table = df_pa.pivot_table(values='Nbr Units', index='User', columns='Location', aggfunc='sum', margins=True, margins_name='Total')
            pivot_table_paah = pivot_table.apply(pd.to_numeric, errors='coerce').astype('Int64')

            df_pa_result = self.calculate_user_times(df_pa,interval_acceptable_delay[2])
            # print("df_pa_result:",df_pa_result)
            df_pa['Nbr Units'] = pd.to_numeric(df_pa['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_pa_units = df_pa['Nbr Units'].sum()
            # print("total_pa_units: ", total_pa_units)
            df_pa_food['Nbr Units'] = pd.to_numeric(df_pa_food['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_pa_units_food = df_pa_food['Nbr Units'].sum()
            # print("total_pa_units_food: ", total_pa_units_food)

        if len(file_data['Adhoc']) > 0:
            df_ah = pd.read_excel(file_data['Adhoc'][0], usecols=["Date Time", "User", "Container", "Dept", "Nbr Units", "To Location"], dtype=str)
            day_of_week = pd.to_datetime(df_ah.loc[0, 'Date Time'], format='%d/%m/%Y %H:%M').day_name()

            df_ah.columns = df_ah.columns.str.strip()
            df_ah_food = df_ah[df_ah['Dept'].astype(str).str.startswith('0')]
            df_ah = df_ah.drop(df_ah_food.index)
            # print("df_ah_food:",df_ah_food)
            # print("df_ah:",df_ah)

            df_ah['Nbr Units'] = df_ah['Nbr Units'].astype(int)
            pivot_table = df_ah.pivot_table(values='Nbr Units', index='User', aggfunc='sum', margins=True, margins_name='Total')
            pivot_table_ah = pivot_table.apply(pd.to_numeric, errors='coerce').astype('Int64')
            # print(pivot_table_ah)
            pivot_table_ah.rename(columns={'Nbr Units': 'Adhoc'}, inplace=True)
            try: # 如果 pivot_table_pa 存在，尝试合并                
                pivot_table_paah = pivot_table_paah.merge(pivot_table_ah, left_index=True, right_index=True, how='left', suffixes=('', '_'))
            except NameError: 
                pass

            df_ah_result = self.calculate_user_times(df_ah,interval_acceptable_delay[2])
            # print("df_ah_result:",df_ah_result)
            df_ah['Nbr Units'] = pd.to_numeric(df_ah['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_ah_units = df_ah['Nbr Units'].sum()
            # print("total_ah_units: ", total_ah_units)
            df_ah_food['Nbr Units'] = pd.to_numeric(df_ah_food['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_ah_units_food = df_ah_food['Nbr Units'].sum()
            # print("total_ah_units_food: ", total_ah_units_food)

        if len(file_data['Rework']) > 0:
            df_rework = pd.read_excel(file_data['Rework'][0], usecols=["Date Time", "User", "Container","Dept","Nbr Units"],dtype=str)
            day_of_week = pd.to_datetime(df_rework.loc[0, 'Date Time'], format='%d/%m/%Y %H:%M').day_name()

            df_rework.columns = df_rework.columns.str.strip()
            # print("df_rework:",df_rework)

            df_rework_result = self.calculate_user_times(df_rework,interval_acceptable_delay[3])
            # print("df_rework_result:",df_rework_result)
            df_rework['Nbr Units'] = pd.to_numeric(df_rework['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_rework_units = df_rework['Nbr Units'].sum()
            # print("total_rework_units: ", total_rework_units)

        if len(file_data['Transfer / Debit']) > 0:
            df_trfdbt = pd.read_excel(file_data['Transfer / Debit'][0], usecols=["Date Time", "User", "Pkt Ctrl Nbr", "Container","Dept","Nbr Units"],dtype=str)
            day_of_week = pd.to_datetime(df_trfdbt.loc[0, 'Date Time'], format='%d/%m/%Y %H:%M').day_name()

            filter_pktctrlnbr = df_trfdbt['Pkt Ctrl Nbr'].str.contains('[a-zA-Z]', regex=True) # 分组条件：'Container' 列包含字母
            df_dbt = df_trfdbt[filter_pktctrlnbr]
            df_trf = df_trfdbt[~filter_pktctrlnbr]

            df_trf.columns = df_trf.columns.str.strip()
            df_trf_food = df_trf[df_trf['Dept'].astype(str).str.startswith('0')]
            df_trf = df_trf.drop(df_trf_food.index)
            # print("df_trf_food:",df_trf_food)
            # print("df_trf:",df_trf)

            df_trf_result = self.calculate_user_times(df_trf,interval_acceptable_delay[4])
            # print("df_trf_result:",df_trf_result)
            df_trf['Nbr Units'] = pd.to_numeric(df_trf['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_trf_units = df_trf['Nbr Units'].sum()
            # print("total_trf_units: ", total_trf_units)
            df_trf_food['Nbr Units'] = pd.to_numeric(df_trf_food['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_trf_units_food = df_trf_food['Nbr Units'].sum()
            # print("total_trf_units_food: ", total_trf_units_food)

            df_dbt.columns = df_dbt.columns.str.strip()
            df_dbt_food = df_dbt[df_dbt['Dept'].astype(str).str.startswith('0')]
            df_dbt = df_dbt.drop(df_dbt_food.index)
            # print("df_dbt_food:",df_dbt_food)
            # print("df_dbt:",df_dbt)
            
            df_dbt_result = self.calculate_user_times(df_dbt,interval_acceptable_delay[4])
            # print("df_dbt_result:",df_dbt_result)
            df_dbt['Nbr Units'] = pd.to_numeric(df_dbt['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_dbt_units = df_dbt['Nbr Units'].sum()
            # print("total_dbt_units: ", total_dbt_units)
            df_dbt_food['Nbr Units'] = pd.to_numeric(df_dbt_food['Nbr Units'], errors='coerce').fillna(0).astype(int)
            total_dbt_units_food = df_dbt_food['Nbr Units'].sum()
            # print("total_dbt_units_food: ", total_dbt_units_food)      

        if len(file_data['ASN verified']) > 0:
            df_asn = pd.read_excel(file_data['ASN verified'][0], usecols=["Last Modified Date", "User"],dtype=str)
            day_of_week = pd.to_datetime(df_asn.loc[0, 'Last Modified Date'], format='%d/%m/%Y %H:%M').day_name()
            df_asn.columns = ["Date Time", "User"]
            df_asn.columns = df_asn.columns.str.strip()
            # print("df_asn:",df_asn)

            df_asn_result = self.calculate_user_times(df_asn,interval_acceptable_delay[0])
            # print("df_asn_result:",df_asn_result)
            total_asn = df_asn.iloc[:, 0].nunique()
            # print(total_asn)

        if len(file_data['Cycle Count']) > 0:
            df_cc = pd.concat([pd.read_excel(file, usecols=["Date Time", "User", "Container", "Dept", "Nbr Units"], dtype=str) for file in file_data['Cycle Count']], ignore_index=True).drop_duplicates()
            day_of_week = pd.to_datetime(df_cc.loc[0, 'Date Time'], format='%d/%m/%Y %H:%M').day_name()

            df_cc.columns = df_cc.columns.str.strip()
            # print("df_cc:",df_cc)
            df_cc_result = self.calculate_user_times(df_cc,interval_acceptable_delay[5])
            # print("df_cc_result:",df_cc_result)


        all_users = list(set(df_receiving_result) | set(df_pa_result) | set(df_ah_result) | set(df_rework_result) | set(df_trf_result) | set(df_dbt_result) | set(df_cc_result) | set(df_asn_result))

        # 创建DataFrame
        df = pd.DataFrame(index=all_users)
        df['Staff'] = [None] * len(df)
        df['ASN'] = df.index.map(df_asn_result.get)
        df['Receiving'] = df.index.map(df_receiving_result.get)
        df['Putaway'] = df.index.map(df_pa_result.get)
        df['Adhoc'] = df.index.map(df_ah_result.get)
        df['Rework'] = df.index.map(df_rework_result.get)
        df['Transfer'] = df.index.map(df_trf_result.get)
        df['Debit'] = df.index.map(df_dbt_result.get)
        df['Cycle Count'] = df.index.map(df_cc_result.get)
        df['Active'] = [None] * len(df)
        df['Shift'] = [None] * len(df)
        df['Idle'] = [None] * len(df)
        df['Tasks'] = [None] * len(df)
        df['Active'] = df[['ASN', 'Receiving', 'Putaway', 'Adhoc', 'Rework', 'Transfer', 'Debit', 'Cycle Count']].sum(axis=1)

        if len(file_data['Inbound Daily Plan']) > 0 and day_of_week != None:
            df_plan = pd.read_excel(file_data['Inbound Daily Plan'][0], day_of_week, dtype=str, header=None)
            df_plan_long = pd.melt(df_plan.reset_index(), id_vars=['index'], var_name='Column', value_name='Value')

            for index, row in df.iterrows():
                # 检查是否存在于 df_plan 的任意位置
                matches = df_plan_long[df_plan_long['Value'] == row.name]  # 假设 df 的索引是您想匹配的值
                if not matches.empty:
                    first_match_index = matches.index[0]
                    actual_row_index = matches.at[first_match_index, 'index']  # 获取原始 df_plan 中的行号
                    column_name = matches.at[first_match_index, 'Column']  # 获取原始 df_plan 中的列号
                    actual_column_index = int(column_name)  # 转换列名为整数索引
                    # print(f"Match found: Row {actual_row_index}, Column {actual_column_index}, Value {row.name}")
                    df.at[index, 'Staff'] = df_plan.iat[actual_row_index, actual_column_index - 1]
                    shift_column_index_time = actual_column_index + 3 if str(row.name).startswith('AGENT') else actual_column_index + 8
                    df.at[index, 'Shift'] = round(float(df_plan.iat[actual_row_index, shift_column_index_time]), 2)
                    shift_column_index_tasks = actual_column_index + 5 if str(row.name).startswith('AGENT') else actual_column_index + 1
                    df.at[index, 'Tasks'] = df_plan.iat[actual_row_index, shift_column_index_tasks]
                    shift_column_index_tasks_2 = actual_column_index + 9 if not str(row.name).startswith('AGENT') else None
                    # print(shift_column_index_tasks_2)
                    if shift_column_index_tasks_2 is not None:
                        value_to_assign = df_plan.iat[actual_row_index, shift_column_index_tasks_2]
                        if pd.notna(value_to_assign):  # 检查值是否非空
                            df.at[index, 'Tasks'] = value_to_assign

        df['Shift'] = pd.to_numeric(df['Shift'], errors='coerce').fillna(0)
        df['Idle'] = df['Active'] - df['Shift']
        df.loc['Active Time Total'] = df[['ASN', 'Receiving', 'Putaway', 'Adhoc', 'Rework', 'Transfer', 'Debit', 'Cycle Count', 'Active', 'Shift', 'Idle']].sum()

        totl_units_data = {'ASN': total_asn, 'Receiving': total_receiving_units,'Putaway': total_pa_units,'Adhoc': total_ah_units,'Rework': total_rework_units,'Transfer': total_trf_units,'Debit': total_dbt_units}
        df.loc['Units Total'] = totl_units_data

        totl_cases_data = {'Receiving': total_receiving_cases}
        df.loc['Cases Received'] = totl_cases_data

        totl_units_data_food = {'Receiving': total_receiving_units_food,'Putaway': total_pa_units_food,'Adhoc': total_ah_units_food, 'Transfer': total_trf_units_food,'Debit': total_dbt_units_food}
        df.loc['Food Units'] = totl_units_data_food

        columns = ['ASN', 'Receiving', 'Putaway', 'Adhoc', 'Rework', 'Transfer', 'Debit', 'Cycle Count']
        indices = [0, 1, 2, 2, 3, 4, 4, 5]
        delay_values = {col: str(interval_acceptable_delay[idx]) + ' min' for col, idx in zip(columns, indices)}
        df.loc['Continuous Delay'] = delay_values

        total_rows = df.loc[['Units Total', 'Cases Received', 'Food Units', 'Continuous Delay', 'Active Time Total']]
        df = df.drop(['Units Total', 'Cases Received', 'Food Units', 'Continuous Delay', 'Active Time Total'])
        df = df.sort_index()
        df = pd.concat([total_rows, df])

        try:              
            pivot_table_paahat = pivot_table_paah.merge(df[['Putaway']], left_index=True, right_index=True, how='left')
            pivot_table_paahat.rename(columns={'Total': 'Units'}, inplace=True)
            pivot_table_paahat.rename(columns={'Putaway': 'Hours'}, inplace=True)
            pivot_table_paahat.loc['Total', 'Hours'] = pivot_table_paahat['Hours'].sum()
            pivot_table_paahat['U / H'] = (pivot_table_paahat['Units'].astype(float) / pivot_table_paahat['Hours'].astype(float)).round(2)
        except NameError: 
            pass
        
        try:
            df_received_dept = pd.read_excel(file_data['Received by Group'][0], usecols=["Inbound Shipment", "Units Received", "Dept"],dtype=str)
            df_received_dept['Group'] = df_received_dept['Dept'].astype(str).str[0].astype(int)
            df_received_dept.loc[df_received_dept['Dept'] == 9868, 'Group'] = 4
            df_received_dept.loc[df_received_dept['Group'].isin([7, 9]), 'Group'] = 6
            df_received_dept = df_received_dept.drop_duplicates(subset=["Inbound Shipment", "Units Received", "Group"])
            df_received_dept['Units Received'] = df_received_dept['Units Received'].astype(int)
            dept_pivot_table = df_received_dept.pivot_table(values='Units Received', index='Group', aggfunc='sum', margins=True, margins_name='Total')
            # print(dept_pivot_table)
        except: 
            pass

        date_today = datetime.now().strftime("%Y%m%d")
        file_name = f"Daily_Inbound_Report_{date_today}.xlsx"

        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Work Summary')
            if 'dept_pivot_table' in locals():
                dept_pivot_table.to_excel(writer, sheet_name='Received Units by Groups')
            if 'df_receiving_food_maybe' in locals():
                df_receiving_food_maybe.to_excel(writer, sheet_name='Unconfirmed Food Receiving')
            if 'pivot_table_paahat' in locals():
                pivot_table_paahat.to_excel(writer, sheet_name='Putaway Productivity')

        wb = load_workbook(file_name)
        # 设置字体和边框样式
        font = Font(name='Arial', size=12)
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))

        # 应用样式到单元格
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.font = font
                    cell.border = thin_border

        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        for sheet in wb.worksheets:
            if sheet.title == 'Work Summary':
                # 找到'Idle'列的索引
                idle_col_index = None
                for col in sheet.iter_cols(1, sheet.max_column):
                    if col[0].value == 'Idle':
                        idle_col_index = col[0].column
                        break
                
                # 如果找到了'Idle'列，应用条件格式
                if idle_col_index:
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=idle_col_index, max_col=idle_col_index):
                        for cell in row:
                            # 检查单元格是否有数值，并根据值设置背景颜色
                            if cell.value is not None:
                                if cell.value < 0:
                                    cell.fill = yellow_fill
                                elif cell.value > 0:
                                    cell.fill = green_fill
                                # 如果值为0或其他非数值数据，则不做任何背景颜色变更
                            else:
                                # 如果是None或非数值，保持默认无填充
                                continue

        # 保存更改
        wb.save(file_name)

        webbrowser.open('file://' + os.path.realpath(file_name))

if __name__ == "__main__":
    app = IDR_1()  # 直接创建 IDR_1 实例，这将是主窗口
    app.mainloop()  # 启动事件循环
