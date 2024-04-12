import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import os
import webbrowser
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import csv
from collections import Counter
import re
from datetime import datetime
import numpy as np
import calendar
import base64
from pathlib import Path
import pyautogui
import pygetwindow as gw
import time
import keyboard
import threading
import xlrd
import tempfile


# 定义一个预定义的窗口类
class PredefinedWindow(tk.Toplevel):
    def __init__(self, parent, title, window_size):
        super().__init__(parent)
        self.title(title)
        self.geometry(window_size)

# 修改MainWindow类，继承自tk.Tk不变，因为它是主窗口
class MainWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Automation Toolkit")
        self.geometry("400x600")  # 设置窗口大小

        # 新增按钮用于打开 Shipment Verify-able Finder 窗口
        self.open_svf_button = ttk.Button(self, text="Shipment Verify Finder", command=self.open_svf_subwindow)
        self.open_svf_button.pack(pady=20)

        # Inside the MainWindow.__init__ method, add the following:
        self.open_rsv_button = ttk.Button(self, text="Receiving", command=self.open_rsv_subwindow)
        self.open_rsv_button.pack(pady=20)

        self.open_subwindow_button = ttk.Button(self, text="Cycle Count", command=self.open_subwindow)
        self.open_subwindow_button.pack(pady=20)

        self.open_rsv_button = ttk.Button(self, text="Food Expiry Report Generator", command=self.open_ferg_subwindow)
        self.open_rsv_button.pack(pady=20)

        # Button to open the second subwindow for SIM barcode typing
        self.open_sim_bc_button = ttk.Button(self, text="SIM Barcode Shooter", command=self.open_sim_bc_subwindow)
        self.open_sim_bc_button.pack(pady=20)

        # 添加“关于”按钮
        self.about_button = ttk.Button(self, text="About", command=self.open_about_window)
        self.about_button.pack(side="bottom", pady=20)

    def open_about_window(self):
        # 创建一个新的顶级窗口来显示关于信息
        about_window = tk.Toplevel(self)
        about_window.title("About")
        # 移除geometry设置以使窗口大小自适应内容

        # 添加普通文本标签
        info_label = tk.Label(about_window, text="To format your data, get the updated version, or find related information, go to:")
        info_label.pack(pady=(10,0))

        # 添加GitHub链接标签，并绑定点击事件
        github_label = tk.Label(about_window, text="https://github.com/Ramielo/Python-Games/releases", fg="blue", cursor="hand2")
        github_label.pack()
        github_label.bind("<Button-1>", lambda e: webbrowser.open("https://github.com/Ramielo/Python-Games/releases"))

        # 添加普通文本标签
        contact_info_label = tk.Label(about_window, text="Contact developer:")
        contact_info_label.pack(pady=(10,0))

        # 添加电子邮件地址标签，并绑定点击事件
        contact_label = tk.Label(about_window, text="charles.liang@davidjones.com.au", fg="blue", cursor="hand2")
        contact_label.pack()
        contact_label.bind("<Button-1>", lambda e: webbrowser.open("mailto:charles.liang@davidjones.com.au"))

    def open_rsv_subwindow(self):
        subwindow = RSV_1(self)
        self.iconify()

        # 子窗口关闭时销毁子窗口并呼出主窗口
        subwindow.protocol("WM_DELETE_WINDOW", lambda: (subwindow.destroy(), self.deiconify()))

    def open_subwindow(self):
        subwindow = Empty_CC_1(self, "800x600")
        # subwindow.grab_set()
        self.iconify()

        # 子窗口关闭时销毁子窗口并呼出主窗口
        subwindow.protocol("WM_DELETE_WINDOW", lambda: (subwindow.destroy(), self.deiconify()))

    def open_sim_bc_subwindow(self):
        subwindow = SIM_BC_1(self)
        self.iconify()

        # 子窗口关闭时销毁子窗口并呼出主窗口
        subwindow.protocol("WM_DELETE_WINDOW", lambda: (subwindow.destroy(), self.deiconify()))

    def open_svf_subwindow(self):
        # 实例化 SVF_1 类
        subwindow = SVF_1(self)
        self.iconify()
        subwindow.protocol("WM_DELETE_WINDOW", lambda: (subwindow.destroy(), self.deiconify()))

    # Then, define the method that this button calls:
    def open_ferg_subwindow(self):
        subwindow = FERG_1(self)
        self.iconify()
        subwindow.protocol("WM_DELETE_WINDOW", lambda: (subwindow.destroy(), self.deiconify()))

class RSV_1(PredefinedWindow):
    def __init__(self, parent, window_size="800x600"):
        super().__init__(parent, "Receiving", window_size)
        self.running = True  # 线程控制标志

        self.current_barcode_index = 0
        self.formatted_barcodes = []

        self.instructions_text = tk.Text(self, wrap="word", bg="white", height=10)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.instructions_text.yview)
        self.instructions_text.configure(yscrollcommand=self.scrollbar.set)

        instructions = """Enter a list of SSCC in the text box below and click [Start receiving].
        
You should run this tool under: [WMS PROD RF][1. DJ Inbound], not under any of its subdirectories.\n"""

        self.instructions_text.insert("1.0", instructions)
        self.instructions_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=0, column=1, sticky="ns", pady=10)
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.data_input_text = tk.Text(self, wrap="word", bg="lightgray", height=15)
        self.data_input_scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.data_input_text.yview)
        self.data_input_text.configure(yscrollcommand=self.data_input_scrollbar.set)
        self.data_input_text.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.data_input_scrollbar.grid(row=1, column=1, sticky="ns", pady=10)
        self.grid_rowconfigure(1, weight=2)
        self.grid_columnconfigure(0, weight=1)

        self.generate_script_button = ttk.Button(self, text="Start receiving", command=self.generate_script)
        self.generate_script_button.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")

        self.protocol("WM_DELETE_WINDOW", self.on_window_close)  # 绑定窗口关闭事件

    def generate_script(self):
        # 从文本框获取数据并检查是否为空
        input_text = self.data_input_text.get("1.0", "end-1c").strip()
        if not input_text:
            # 直接在instructions_text文本框中显示警告信息
            self.instructions_text.insert("end", "\nNo SSCC list specified, please check the input box below.\n")
            self.instructions_text.see("end")  # 确保警告信息滚动到可视区域
            return  # 结束方法的执行
        
        # 如果输入框不为空，则继续处理数据
        self.barcodes = input_text.split("\n")
        self.formatted_barcodes = ["0" * (13 - len(barcode)) + barcode if len(barcode) < 13 else barcode for barcode in self.barcodes if barcode.strip()]
        
        # 锁定按钮和文本框，然后开始线程
        self.generate_script_button.config(state='disabled', text='Executing')
        self.data_input_text.config(state='disabled')
        threading.Thread(target=self.input_barcodes).start()

    def input_barcodes(self):
        time.sleep(1)
        if not self.activate_notepad_plus_plus():
            self.instructions_text.insert("end", "\nNo running [vipdjwmsapp.davidjones.com.au - PuTTY] window detected.\n")
            self.instructions_text.see("end")
            self.reset_button_state()
            return

        while self.current_barcode_index < len(self.formatted_barcodes) and self.running:
            barcode_to_send = self.formatted_barcodes[self.current_barcode_index]
            pyautogui.typewrite('1' + '\n')
            time.sleep(1)
            pyautogui.typewrite(barcode_to_send + '\n')
            time.sleep(5)
            pyautogui.hotkey('ctrl', 'a')
            time.sleep(1)
            pyautogui.hotkey('ctrl', 'x')
            time.sleep(1)

            if not self.is_notepad_plus_plus_active():
                self.instructions_text.insert("end", f"\n[vipdjwmsapp.davidjones.com.au - PuTTY] window is deactivated or closed, execution proceeded up to SSCC: {barcode_to_send}.\n")
                self.instructions_text.see("end")
                self.current_barcode_index += 1
                break
            self.current_barcode_index += 1

        self.reset_button_state()
        if self.current_barcode_index >= len(self.formatted_barcodes):
            self.instructions_text.insert("end", "\nAll SSCC receiving completed.\n")
            self.instructions_text.see("end")
            self.current_barcode_index = 0  # 重置当前处理的条形码索引
            self.data_input_text.config(state='normal')

    def is_notepad_plus_plus_active(self):
        # 这个方法需要根据实际情况来实现，以检查vipdjwmsapp.davidjones.com.au - PuTTY是否处于激活状态
        try:
            notepad_window = gw.getWindowsWithTitle("vipdjwmsapp.davidjones.com.au - PuTTY")[0]
            return notepad_window.isActive
        except IndexError:
            return False

    def activate_notepad_plus_plus(self):
        try:
            notepad_window = gw.getWindowsWithTitle("vipdjwmsapp.davidjones.com.au - PuTTY")[0]
            notepad_window.activate()
            time.sleep(1)  # Give the window some time to become active
            return self.is_notepad_plus_plus_active()
        except IndexError:
            return False
        
    def reset_button_state(self):
        self.generate_script_button.config(state='normal', text='Start receiving')

    def on_window_close(self):
        self.running = False  # 通知线程停止
        self.thread.join()  # 等待线程结束
        self.destroy()  # 销毁窗口

class FERG_1(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Food Expiry Report Generator")
        self.geometry("800x600")

        # Text box for guiding the operation and displaying information
        self.instruction_text = tk.Text(self, wrap=tk.WORD, height=20, width=80)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.instruction_text.yview)
        self.instruction_text.configure(yscrollcommand=self.scrollbar.set)
        
        # Layout text box and scrollbar using grid
        self.instruction_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=0, column=1, sticky="ns")

        # Configure grid row/column weights
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Add initial text to the text box
        initial_text = """This tool generates Food Expiry Report.

Retrieve the following data from WMS:

[Pick Location Inquiry]: Set [Department] number to [0].

Import data sheet using the [Import Excel File] button.\n"""
        self.instruction_text.insert(tk.END, initial_text)

        # Button for file selection
        self.file_button = tk.Button(self, text="Import Excel File", command=self.process_file)
        
        # Adjust button position using grid
        self.file_button.grid(row=1, column=0, columnspan=2, pady=10, sticky="ew")

        # Ensure grid rows and columns adjust with the content size
        self.grid_rowconfigure(1, weight=1)  # Set weight for the row with the button
        self.grid_columnconfigure(1, weight=1)  # Set weight for the scrollbar column

    def process_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls")])
        if file_path:
            try:
                self.process_xls(file_path)
                self.instruction_text.see(tk.END)  # 滚动到文本框底部
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def process_xls(self, file_path):
        # Read .xls file using xlrd
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)

        # Process data using Pandas
        df = pd.read_excel(file_path, dtype=str)
        
        # Check for all required columns
        required_columns = ['Location', 'Dept #', 'Barcode', 'Current Qty', 'Expiry Date', 'Brand', 'SKU', 'Description', 'Case Qty']
        for column in required_columns:
            if column not in df.columns:
                raise ValueError(f"Missing required column: {column}")

        df = df[required_columns]
        # Insert two empty columns at the specified position
        df.insert(df.columns.get_loc("Expiry Date") + 1, 'temp1', '')
        df.insert(df.columns.get_loc("Expiry Date") + 2, 'temp2', '')

        # Remove rows that do not meet the condition
        df = df[df['Dept #'].astype(str).str.startswith(('04', '06'))]

        df['Current Qty'] = df['Current Qty'].str.split(' ').str[0].astype(int)

        # 处理 Expiry Date，只保留日期部分
        df['Expiry Date'] = pd.to_datetime(df['Expiry Date'].str.split(' ').str[0], format='%d/%m/%Y')


        # 计算日期与当前日期的差值，保存在temp1里面
        df['temp1'] = (df['Expiry Date'] - pd.to_datetime('today')).dt.days

        # 根据差值更新temp2列
        conditions = [
            df['temp1'] < 30,
            df['temp1'] < 60,
            df['temp1'] < 120
        ]
        choices = ['<30', '<60', '<120']
        df['temp2'] = np.select(conditions, choices, default='')  # 注意：pd.np.select 在未来版本的pandas中可能被替换或更改

        # Convert 'Expiry Date' back to string in the format %d/%m/%Y
        df['Expiry Date'] = df['Expiry Date'].dt.strftime('%d/%m/%Y')
        
        # Sort the DataFrame by 'temp1' in ascending order
        df = df.sort_values(by='temp1', ascending=True)
        
        # After processing data, export to .xlsx format
        self.export_to_xlsx(df)

    def export_to_xlsx(self, df):
        # Prompt user for location to save the processed file
        root = tk.Tk()
        root.withdraw()  # Hide the main window
        export_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if export_file_path:
            # 使用ExcelWriter和openpyxl引擎
            with pd.ExcelWriter(export_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
                
                worksheet = writer.sheets['Sheet1']
                
                # 设置字体和边框样式
                bold_font = Font(name='Arial', size=10, bold=True)
                thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))

                # 为 '<30'、'<60' 和 '<120' 设置反色字体和粗体
                black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
                white_font_bold = Font(color='FFFFFF', bold=True)
                orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
                dark_blue_font_bold = Font(color='8B0000', bold=True) # 深蓝色作为橙色的高对比色
                yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                black_font_bold = Font(color='A52A2A', bold=True)   
               
                # 合并列标题
                worksheet.merge_cells('F1:G1')
                worksheet.cell(row=1, column=6).value = 'Date to Expiry'
                
                # 条件格式
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet['G' + str(row)]  # G列
                    if cell.value == '<30':
                        cell.font = white_font_bold
                        cell.fill = black_fill
                    elif cell.value == '<60':
                        cell.font = dark_blue_font_bold
                        cell.fill = orange_fill
                    elif cell.value == '<120':
                        cell.font = black_font_bold
                        cell.fill = yellow_fill

                # 应用字体和边框
                for row in worksheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
                    for cell in row:
                        cell.font = Font(name='Arial', size=10, bold=cell.font.bold, color=cell.font.color)
                        cell.border = thin_border
                        if cell.row == 1:  # 标题行
                            cell.font = bold_font

                # 列宽自动适应内容
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

            # UI反馈
            self.instruction_text.insert(tk.END, f"\nProcessing complete. File has been saved to: {export_file_path}\n")

            try:
                if os.name == 'nt':  # 如果是Windows系统
                    os.startfile(export_file_path)
                else:  # 非Windows系统
                    webbrowser.open(export_file_path)
            except Exception as e:
                self.instruction_text.insert(tk.END, f"\nFailed to open the file: {str(e)}\n")


class SVF_1(PredefinedWindow):
    def __init__(self, parent, window_size="800x600"):
        super().__init__(parent, "Shipment Verify-able Finder", window_size)

        # Create a textbox with a scrollbar for instructions
        self.instructions_text = tk.Text(self, wrap="word", bg="white", height=10)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.instructions_text.yview)
        self.instructions_text.configure(yscrollcommand=self.scrollbar.set)

        # Insert guidance text
        instructions = """This tool automatically identifies shipments ready for verification.

Retrieve the following data from WMS:

[Inbound Shipment Inquiry]: Set [Status From] and [Status To] to [In receiving]. 

[Case Inquiry]: Set [From Status] and [To Status] to [Consumed]. Use [More Criteria] to set [From Consume Priority Date].

Import all data sheets at once using the [Import Excel Files] function.

Please consult the release notes for data format and additional information.\n"""

        self.instructions_text.insert("1.0", instructions)

        # Layout the textbox and scrollbar
        self.instructions_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=0, column=1, sticky="ns", pady=10)

        # Configure grid row/column weights to ensure the textbox expands with the window size
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # Create the "Import Excel Files" button and layout using grid
        self.import_button = ttk.Button(self, text="Import Excel Files", command=self.import_excel)
        self.import_button.grid(row=1, column=0, columnspan=2, pady=10, sticky="ew")
        self.data_dict = {}

    def import_excel(self):
        # Clear the textbox and display the starting message
        self.instructions_text.insert(tk.END, "\nStarting to import data...\n")
        self.update_idletasks()  # Update UI

        # Open file dialog to select Excel files
        file_paths = filedialog.askopenfilenames(filetypes=[("Excel 97-2003 Workbook", "*.xls")])
        # 预分类文件
        files_cases_consumed = []
        files_cases_virtual = []

        for file_path in file_paths:
            try:
                df = pd.read_excel(file_path, nrows=2)  # 只加载前两行进行检查
                if 'Consumed' in df.iloc[1].to_string():
                    files_cases_consumed.append(file_path)
                    print("files_cases_consumed",files_cases_consumed)
                elif 'In Inventory, Not Putaway' in df.iloc[1].to_string():
                    files_cases_virtual.append(file_path)
                    print("files_cases_virtual",files_cases_virtual)
            except Exception as e:
                print(f"Error previewing file {file_path}: {e}")

        case_inquiry_data = {}
        Inbound_Shipment_inquiry_data = {}

        inbound_shipment_inquiry_files = [file for file in file_paths if "Inbound Shipment Inquiry" in file]
        if len(inbound_shipment_inquiry_files) > 1:
            self.instructions_text.insert(tk.END, "\nError: More than one 'Inbound Shipment Inquiry' file found.\n")
            return
        elif len(inbound_shipment_inquiry_files) == 1:
            try:
                df_inbound = pd.read_excel(inbound_shipment_inquiry_files[0], usecols=["Inbound Shipment", "Cases Shipped", "Cases Received"])
                print (df_inbound)
                Inbound_Shipment_inquiry_data = df_inbound.to_dict(orient='list')
                # print(Inbound_Shipment_inquiry_data)
            except Exception as e:
                self.instructions_text.insert(tk.END, f"\nError processing file {inbound_shipment_inquiry_files[0]}: {e}\n")

        for file_path in files_cases_consumed:
            try:
                df = pd.read_excel(file_path) # 依次处理cases_consumed表
                df = df.drop_duplicates(subset=['Case']) # 按Case number去重
                df = df[df['Case'].astype(str).str.match(r'^\d{10,}$')] # 仅保留Case列=[长度至少为10位的纯数字字符串]的行

                shipment_col = None
                for col in df.columns:
                    if "shipment" in col.lower() or "shpmt" in col.lower():
                        shipment_col = col # 把cases_consumed表的“Inbound Shipment”列取出来
                        break
                if shipment_col is None:
                    raise ValueError(f"Shipment column not found in {file_path}")
                    
                df[shipment_col] = df[shipment_col].apply(lambda x: str(int(x)) if pd.notnull(x) and float(x).is_integer() else x) # “Inbound Shipment”列规整格式：整数形式的字符串
                df.dropna(subset=[shipment_col], inplace=True) # “Inbound Shipment”列删空行
               
                counts = df[shipment_col].value_counts().to_dict()
                print (counts)

                for key, value in counts.items():
                    if key in case_inquiry_data:
                        case_inquiry_data[key] += value
                    else:
                        case_inquiry_data[key] = value

            except Exception as e:
                self.instructions_text.insert(tk.END, f"\nError processing file {file_path}: {e}\n")

        for file_path in files_cases_virtual:
            if "Case Inquiry" in file_path:
                try:
                    df = pd.read_excel(file_path)
                    
                    # 确定“Inbound Shipment”列
                    shipment_col = None
                    for col in df.columns:
                        if "shipment" in col.lower() or "shpmt" in col.lower():
                            shipment_col = col
                            break

                    if shipment_col is None:
                        raise ValueError(f"Shipment column not found in {file_path}")
                    

                    df = df.drop_duplicates(subset=['Case'])
                    # 筛选出“Case”列中符合新格式要求的行：包含数字或字母，长度不足10
                    df = df[df['Case'].astype(str).str.match(r'^[a-zA-Z0-9]{1,9}$')]
                        
                        
                    df[shipment_col] = df[shipment_col].apply(lambda x: str(int(x)) if pd.notnull(x) and float(x).is_integer() else x)
                    df.dropna(subset=[shipment_col], inplace=True)

                    counts = df[shipment_col].value_counts().to_dict()

                    for key, value in counts.items():
                        value_str = str([value])
                        if key in case_inquiry_data:
                            # 如果key已存在于case_inquiry_data中
                            # 直接将原有值和新的格式化字符串值组成一个列表
                            case_inquiry_data[key] = [case_inquiry_data[key], value_str]
                        else:
                            # 如果键不存在，则以字符形式（值用方括号包围）添加键值对
                            case_inquiry_data[key] = value_str
                

                except Exception as e:
                    self.instructions_text.insert(tk.END, f"\nError processing file {file_path}: {e}\n")


        # Add the fourth column for matches with Case Inquiry data
        if Inbound_Shipment_inquiry_data: #Inbound_Shipment_inquiry_data已删，修改
            Inbound_Shipment_inquiry_data['Cases Consumed'] = [
                case_inquiry_data.get(str(shipment), 0) for shipment in Inbound_Shipment_inquiry_data['Inbound Shipment']
            ]

        if len(inbound_shipment_inquiry_files) == 1:
            try:
                # 加载整个文件以查找特定列
                df_full = pd.read_excel(inbound_shipment_inquiry_files[0])
                # 寻找同时包含"first", "date", "time"的列标题
                datetime_col = [col for col in df_full.columns if all(keyword in col.lower() for keyword in ["first", "date", "time"]) or all(keyword in col.lower() for keyword in ["received", "date"])]
                
                if datetime_col:
                    # 假定只有一个符合条件的列，获取其数据
                    datetime_data = df_full[datetime_col[0]].tolist()
                    # 添加到Inbound_Shipment_inquiry_data字典中
                    Inbound_Shipment_inquiry_data['First DateTime'] = datetime_data
                else:
                    # 如果找不到符合条件的列，直接跳过相关操作
                    self.instructions_text.insert(tk.END, "\nNo suitable datetime column found.\n")
                    # 或者在这里添加任何你认为合适的逻辑
            except Exception as e:
                self.instructions_text.insert(tk.END, f"\nError processing additional data from file {inbound_shipment_inquiry_files[0]}: {e}\n")


        if Inbound_Shipment_inquiry_data:
            # 将字典转换为DataFrame
            df_inbound = pd.DataFrame(Inbound_Shipment_inquiry_data)
            # 删除重复行
            df_inbound = df_inbound.drop_duplicates(subset=df_inbound.columns[:3].tolist())
            # 更新Inbound_Shipment_inquiry_data字典，以便于展示和之后的处理
            Inbound_Shipment_inquiry_data = df_inbound.to_dict(orient='list')

        self.instructions_text.insert(tk.END, "\nData import completed.\n")

        # 在你的方法最末尾加入以下代码，紧接在展示结果到GUI的代码之后

        def summarize_and_write_to_txt(Inbound_Shipment_inquiry_data, case_inquiry_data):
            # Convert dictionary to DataFrame
            df = pd.DataFrame(Inbound_Shipment_inquiry_data)
            print(df)
            # Ensure correct datetime format if 'First DateTime' column exists
            if 'First DateTime' in df.columns:
                df['First DateTime'] = pd.to_datetime(df['First DateTime'], format='%d/%m/%Y %H:%M')
                # Calculate today's date
                today = pd.to_datetime("today")
                # Shipments that are about to expire
                upcoming_expired_shipments = df[df['First DateTime'] < (today - pd.Timedelta(days=1))]['Inbound Shipment']
            else:
                upcoming_expired_shipments = pd.Series([], dtype='object')  # Empty series if 'First DateTime' does not exist

            matched_case_count_pairs = {}

            # 遍历DataFrame，检查并处理Cases Consumed字段
            for index, row in df.iterrows():
                matched_case_count = row['Cases Consumed']
                if isinstance(matched_case_count, list) and len(matched_case_count) == 2:
                    # 存储Inbound Shipment和B的键值对
                    matched_case_count_pairs[row['Inbound Shipment']] = matched_case_count[1]
                    # 将Cases Consumed中的A替换原来的列表位置
                    df.at[index, 'Cases Consumed'] = matched_case_count[0]

            # 现在，使用更新后的DataFrame执行原有的逻辑
            verified_shipments = df[(df['Cases Shipped'] == df['Cases Received']) & (df['Cases Received'] == df['Cases Consumed'])]['Inbound Shipment']
            partial_verified_shipments = df[(df['Cases Received'] == df['Cases Consumed']) & (df['Cases Received'] < df['Cases Shipped'])]['Inbound Shipment']

            
            # Write summary to TXT file
            txt_file_path = 'shipment_summary.txt'
            full_path = os.path.abspath(txt_file_path)  # 转换为绝对路径
            with open(txt_file_path, 'w', encoding='utf-8') as file:
                # 写入可以被确认的进货信息
                file.write("Shipments that can be VERIFIED:\n")
                for shipment in verified_shipments:
                    # 检查是否需要添加虚拟案例的注释
                    if shipment in matched_case_count_pairs:
                        file.write(f"{shipment} (Virtual case detected: {matched_case_count_pairs[shipment]})\n")
                    else:
                        file.write(f"{shipment}\n")
                
                # 写入可以被部分确认的进货信息
                file.write("\nShipments that can be verified but not all have been received:\n")
                for shipment in partial_verified_shipments:
                    # 检查是否需要添加虚拟案例的注释
                    if shipment in matched_case_count_pairs:
                        file.write(f"{shipment} (virtual case detected: {matched_case_count_pairs[shipment]})\n")
                    else:
                        file.write(f"{shipment}\n")

                if 'First DateTime' in df.columns:  # Write this section only if 'First DateTime' exists
                    file.write("\nShipments that are about to expire:\n")
                    file.write('\n'.join(upcoming_expired_shipments.astype(str)) + '\n\n')

                # Additional Section: Insert Case Inquiry and Inbound Shipment Inquiry Data summaries
                file.write("\nInbound Shipment Inquiry Data Summary:\n")
                if 'First DateTime' in Inbound_Shipment_inquiry_data:
                    for i in range(len(Inbound_Shipment_inquiry_data['Inbound Shipment'])):
                        text_line = f"{Inbound_Shipment_inquiry_data['Inbound Shipment'][i]}, {Inbound_Shipment_inquiry_data['Cases Shipped'][i]}, {Inbound_Shipment_inquiry_data['Cases Received'][i]}, {Inbound_Shipment_inquiry_data['Cases Consumed'][i]}"
                        if 'First DateTime' in Inbound_Shipment_inquiry_data:
                            text_line += f", {Inbound_Shipment_inquiry_data['First DateTime'][i]}"
                        file.write(text_line + "\n")

            # Open the TXT file automatically after saving
            webbrowser.open(txt_file_path)

            # 在函数最后返回txt文件路径
            return full_path


        if Inbound_Shipment_inquiry_data:
            # 调用函数并接收返回的TXT文件路径
            full_txt_file_path = summarize_and_write_to_txt(Inbound_Shipment_inquiry_data, case_inquiry_data)

            # 在文本框中显示TXT文件的保存路径和提示信息
            verification_summary_message = f"\nVerification summary has been saved to:\n{full_txt_file_path}\n"
            self.instructions_text.insert(tk.END, verification_summary_message)
        self.instructions_text.see(tk.END)



class SIM_BC_1(PredefinedWindow):
    def __init__(self, parent, window_size="800x600"):
        super().__init__(parent, "SIM Barcode Shooter", window_size)
        
        self.instructions_text = tk.Text(self, wrap="word", bg="white", height=10)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.instructions_text.yview)
        self.instructions_text.configure(yscrollcommand=self.scrollbar.set)
        
        instructions = """Enter a list of barcodes in the text box below and click "Load Barcodes".
        
This tool is useful for entering a large number of different barcodes at once in SIM, especially suitable for creating Claims or Transfers.\n"""
        
        self.instructions_text.insert("1.0", instructions)
        
        self.instructions_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=0, column=1, sticky="ns", pady=10)
        
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        
        self.data_input_text = tk.Text(self, wrap="word", bg="lightgray", height=15)
        self.data_input_scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.data_input_text.yview)
        self.data_input_text.configure(yscrollcommand=self.data_input_scrollbar.set)
        self.data_input_text.grid(row=1, column=0, sticky="nsew", padx=10, pady=10)
        self.data_input_scrollbar.grid(row=1, column=1, sticky="ns", pady=10)
        
        self.grid_rowconfigure(1, weight=2)
        self.grid_columnconfigure(0, weight=1)
        
        self.generate_script_button = ttk.Button(self, text="Load Barcodes", command=self.generate_script)
        self.generate_script_button.grid(row=2, column=0, columnspan=2, pady=10, sticky="ew")

    def generate_script(self):
        barcodes = self.data_input_text.get("1.0", "end-1c").split("\n")
        
        # 检查所有输入是否为数字
        if not all(barcode.isdigit() for barcode in barcodes if barcode.strip()):
            error_message = "\nInput error: Please ensure all inputs are numbers.\n"
            self.instructions_text.insert("end", error_message)
            self.instructions_text.see("end")
            return
        
        # 格式化条形码：确保每个条形码长度为13位
        formatted_barcodes = ["0"*(13-len(barcode)) + barcode if len(barcode) < 13 else barcode for barcode in barcodes if barcode.strip()]
        
        threading.Thread(target=self.wait_and_input_barcodes, args=(formatted_barcodes,)).start()

    def wait_and_input_barcodes(self, formatted_barcodes):
        self.instructions_text.insert("end", "\nReady. Press [Ctrl+Alt+F] to shoot barcodes.\n")
        self.instructions_text.see("end")
        self.generate_script_button.config(state='disabled')
        
        # 等待用户按下WIN+N
        keyboard.wait('ctrl+alt+f')
        # 在开始输入条形码之前添加一个短暂的延迟
        time.sleep(1)
        for barcode_to_send in formatted_barcodes:
            pyautogui.typewrite(barcode_to_send)  # 确保包括了按回车
            time.sleep(0.5)
        
        # 注意：更新 GUI 的操作需要在主线程中执行
        # 在这里简化为直接调用，但在实际应用中可能需要使用线程安全的方式更新 GUI
        self.instructions_text.insert("end", "\nMission completed, waiting for reload.\n")
        self.instructions_text.see("end")
        self.generate_script_button.config(state='normal')

class Empty_CC_1(PredefinedWindow):
    def __init__(self, parent, window_size):
        super().__init__(parent, "Cycle Count", window_size)
        self.combined_df = None  # 初始化combined_df属性

        # 创建带滚动条的文本框
        self.instructions_text = tk.Text(self, wrap="word", bg="white", height=10)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.instructions_text.yview)
        self.instructions_text.configure(yscrollcommand=self.scrollbar.set)

        # 插入指导步骤文本
        instructions = """This tool automates [Cycle Count].

Retrieve the following data from WMS:

[Pick Location Inquiry]: Set a [Zone] number (such as 01, 09, 15).

Import all data sheets using the [Import Excel Files] function.

Note: The WMS PROD RF interface may freeze, potentially resulting in missed locations in [Cycle Count] outcomes.

Please consult the release notes for data format and additional information.\n"""
        self.instructions_text.insert("1.0", instructions)

        # 布局文本框和滚动条
        self.instructions_text.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=0, column=1, sticky="ns", pady=10)

        # 配置网格行/列的权重，确保文本框可以随窗口大小调整而扩展
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # 创建单选按钮的选中值变量
        self.cycle_count_option = tk.IntVar()
        self.cycle_count_option.set(1)  # 默认选中第一个选项

        # 创建单选按钮
        ttk.Radiobutton(self, text="Cycle Count Empty", variable=self.cycle_count_option, value=1).grid(row=1, column=0, sticky="w", padx=20)
        ttk.Radiobutton(self, text="Cycle Count Everything", variable=self.cycle_count_option, value=2).grid(row=1, column=0, sticky="e", padx=20)

        # 创建导入按钮，并使用grid布局
        self.import_button = ttk.Button(self, text="Import Excel Files", command=self.import_excel)
        # 适当调整按钮的放置位置
        self.import_button.grid(row=2, column=0, columnspan=2, pady=10, sticky="sn")

        # 添加下一步按钮
        self.next_button = ttk.Button(self, text="Next", state="disabled", command=self.goto_next_step)  # 下一步按钮初始不可用
        self.next_button.grid(row=2, column=1, sticky="se", padx=10, pady=10)

        # 配置网格布局权重，确保按钮固定在底部边缘
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(1, weight=1)

       
    def goto_next_step(self):
        if self.combined_df is not None:
            self.destroy()  # 关闭当前窗口
            next_window = Empty_CC_2(self.master, "800x600", self.combined_df)
            # next_window.grab_set()
            # 子窗口关闭时销毁子窗口并呼出主窗口
            next_window.protocol("WM_DELETE_WINDOW", lambda: (next_window.destroy(), self.master.deiconify()))
        else:
            messagebox.showwarning("Warning", "No data to proceed.")

 
    def import_excel(self):
        self.instructions_text.insert("end", "\nStart to import data...\n")
        self.instructions_text.see("end")  # 确保新消息可见
        
        file_paths = filedialog.askopenfilenames(title="Select Excel Files", filetypes=(("Excel files", "*.xls"),))
        
        if file_paths:
            try:
                new_data_frames = []
                required_columns_base = ["Location", "Current Qty", "Last Count Date"]
                required_columns = required_columns_base + (["Barcode", "Expiry Date"] if self.cycle_count_option.get() == 2 else [])
                
                for file_path in file_paths:
                    df = pd.read_excel(file_path)
                    
                    # 检查是否包含所有必需的列
                    if not all(col in df.columns for col in required_columns):
                        self.show_error_dialog("Critical data missing", "450x150", "Critical data missing in one or more files: Please ensure all imported files contain the following three columns of data:\nLocation, Current Qty, Last Count Date.")
                        return
                    
                    # 根据cycle_count_option的值应用不同的数据处理逻辑
                    if self.cycle_count_option.get() == 1:
                        # 仅保留“Current Qty”为空的行
                        df_filtered = df[required_columns][df["Current Qty"].isna()]
                    else:
                        # cycle_count_option等于2时，不过滤“Current Qty”为空的行，直接使用筛选后的数据
                        df_filtered = df[required_columns]
                    
                    new_data_frames.append(df_filtered)
                
                self.combined_df = pd.concat(new_data_frames).drop_duplicates().reset_index(drop=True)

                # 检查合并后的DataFrame是否为空
                if self.combined_df.empty:
                    self.instructions_text.insert("end", "\nNo empty location can be found.\n")
                else:
                    self.next_button['state'] = 'normal'
                    final_data_message = "\nData import completed, final data has {} rows and {} columns.\n\nClick [Next] to process data.\n".format(self.combined_df.shape[0], self.combined_df.shape[1])
                    self.instructions_text.insert("end", final_data_message)
                self.instructions_text.see("end")  # Auto-scroll to the new text

            except Exception as e:
                print("Unable to read Excel file:", e)
                self.show_error_dialog("Error", "300x100", f"Unable to read Excel file: {e}")


    def show_error_dialog(self, title, size, message):
        error_window = PredefinedWindow(self, title, size)
        message_label = ttk.Label(error_window, text=message)
        message_label.pack(pady=10, padx=10, fill="both", expand=True)
        ttk.Button(error_window, text="Confirm", command=error_window.destroy).pack(pady=(0,10), padx=10)

        def adjust_wraplength(event):
            # 调整wraplength为窗口宽度减去一些像素以留出边距
            message_label.config(wraplength=event.width - 20)

        # 绑定到窗口大小变化事件
        error_window.bind("<Configure>", adjust_wraplength)


class Empty_CC_2(PredefinedWindow):
    def __init__(self, parent, window_size, combined_df):
        super().__init__(parent, "Cycle Count Empty", window_size)
        self.combined_df = combined_df  # 存储传递的DataFrame
        
        self.initialize_ui()
        self.execution_running = False  # 添加一个执行状态标志
        self.current_location_index = 0  # 添加用于跟踪当前位置的属性

    def initialize_ui(self):
        # 创建带滚动条的文本框
        self.text_box = tk.Text(self, wrap="word", bg="white", height=10)
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.text_box.yview)
        self.text_box.configure(yscrollcommand=self.scrollbar.set)

        # 布局文本框和滚动条
        self.text_box.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        self.scrollbar.grid(row=0, column=1, sticky="ns", pady=10)

        # 添加下拉选择框以选择排序方式
        self.sort_option_var = tk.StringVar(self)
        self.sort_option_var.set("Ascending")  # 默认值
        self.sort_options = ["Ascending", "Descending"]
        self.sort_option_menu = ttk.Combobox(self, textvariable=self.sort_option_var, values=self.sort_options, state="readonly")
        self.sort_option_menu.grid(row=1, column=0, padx=10, pady=5, sticky="ew")
        self.sort_option_menu.bind("<<ComboboxSelected>>", self.update_display)

        # 添加日期选择的下拉菜单
        # 首先，将'Last Count Date'列转换为datetime对象
        self.combined_df['Last Count Date'] = pd.to_datetime(self.combined_df['Last Count Date'], format="%d/%m/%Y %H:%M").fillna(pd.Timestamp('1900-01-01'))

        # 最后，提取日期部分。这一步确保所有日期都没有时间部分
        self.combined_df['Last Count Date'] = self.combined_df['Last Count Date'].apply(lambda x: x.date())


        unique_dates = sorted(self.combined_df['Last Count Date'].unique())
        
        # 生成带有前缀和后缀的日期选项
        unique_dates_str = [f"Update to date: {date}" for date in unique_dates]
        if unique_dates_str:
            # 在最后一个日期选项后加上"(All data)"
            unique_dates_str[-1] += " (All data selected)"

        # 设置默认值为最后一天的日期，即所有数据
        self.date_option_var = tk.StringVar(self)
        self.date_option_var.set(unique_dates_str[-1] if unique_dates_str else "")

        self.date_options_menu = ttk.Combobox(self, textvariable=self.date_option_var, values=unique_dates_str, state="readonly")
        self.date_options_menu.grid(row=2, column=0, padx=10, pady=5, sticky="ew")
        self.date_options_menu.bind("<<ComboboxSelected>>", self.update_display)

        # 添加选择 Sleep 时间的提示文字
        self.sleep_time_label = tk.Label(self, text="Please select the script execution speed (sleep time between each step, in milliseconds):")
        self.sleep_time_label.grid(row=3, column=0, padx=10, pady=(10, 0), sticky="nw")

        # 创建 Sleep 时间选项的下拉菜单
        self.sleep_time_var = tk.StringVar(self)
        self.sleep_time_options = ['500', '1000', '1500', '2000', '2500', '3000', '3500', '4000', '4500', '5000']  # 以毫秒为单位
        self.sleep_time_menu = ttk.Combobox(self, textvariable=self.sleep_time_var, values=self.sleep_time_options, state="readonly")
        self.sleep_time_menu.set('1000')  # 设置默认值
        self.sleep_time_menu.grid(row=3, column=0, padx=10, pady=5, sticky="e")
        
        # 确保添加对应的 grid_rowconfigure 调用以适应新组件...

        # 初始数据显示
        self.update_display()

        # 配置网格行/列的权重，确保文本框和下拉框可以随窗口大小调整而扩展
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        # 添加前一步按钮
        self.prev_button = ttk.Button(self, text="Previous", command=self.go_back_to_empty_cc_1)
        self.prev_button.grid(row=3, column=0, sticky="sw", padx=10, pady=10)

        # 添加导出至Autokey脚本文件按钮
        self.export_button = ttk.Button(self, text="Start Execution", command=self.start_export_thread)
        self.export_button.grid(row=3, column=1, sticky="se", padx=10, pady=10)


        # 配置网格布局权重，确保按钮固定在底部边缘
        self.grid_rowconfigure(3, weight=1)
        self.grid_columnconfigure(1, weight=1)

    def start_export_thread(self):
        export_thread = threading.Thread(target=self.export_to_txt, name="ExportThread")
        export_thread.daemon = True  # 设置为守护线程，这样当主程序退出时线程也会被终止
        export_thread.start()

    def update_display(self, event=None):
        # 解析选择的日期
        selected_date_str = self.date_option_var.get().replace("Update to date: ", "").replace(" (All data)", "")
        try:
            selected_date = pd.to_datetime(selected_date_str).date()
        except ValueError:
            selected_date = self.combined_df['Last Count Date'].max()

        ascending = True if self.sort_option_var.get() == "Ascending" else False
        
        # 使用新属性来存储过滤和排序后的DataFrame
        self.filtered_df = self.combined_df[self.combined_df['Last Count Date'] <= selected_date].sort_values(by="Location", ascending=ascending)
        self.filtered_df['Location'] = self.filtered_df['Location'].apply(lambda x: f"{x:09}")
        if 'Barcode' in self.filtered_df.columns:
            # 如果存在，则对 'Barcode' 列应用格式化函数
            self.filtered_df['Barcode'] = self.filtered_df['Barcode'].apply(lambda x: format(x, '.0f') if isinstance(x, (int, float)) else x)
        else:
            pass
        rows_count_info = f"Total rows displayed: {len(self.filtered_df)}\n"
        df_string = self.filtered_df.to_string(index=False)
        display_string = rows_count_info + df_string

        self.text_box.delete("1.0", tk.END)
        self.text_box.insert("1.0", display_string)


    def go_back_to_empty_cc_1(self):
        # 关闭当前窗口
        self.destroy()
        # 直接创建并显示一个Empty_CC_1窗口的实例
        # 注意: 这假设Empty_CC_1类已经在这个文件中被定义或正确导入
        empty_cc_1_window = Empty_CC_1(self.master, "800x600")
        # empty_cc_1_window.grab_set()

    def is_notepad_plus_plus_active(self):
        # 检查vipdjwmsapp.davidjones.com.au - PuTTY窗口是否最上层和未关闭。
        # 返回True如果vipdjwmsapp.davidjones.com.au - PuTTY是活动的且最上层，否则返回False。
        try:
            notepad_window = gw.getWindowsWithTitle("vipdjwmsapp.davidjones.com.au - PuTTY")[0]
            return notepad_window.isActive and not notepad_window.isMinimized
        except IndexError:
            # 如果找不到窗口，认为它已被关闭
            return False


    def export_to_txt(self):

        current_thread = threading.current_thread()
        print(f"This method is running in thread: {current_thread.name}")

        def activate_notepad_plus_plus():
            try:
                notepad_window = gw.getWindowsWithTitle("vipdjwmsapp.davidjones.com.au - PuTTY")[0]
                if not notepad_window.isActive:
                    notepad_window.activate()  # Activate without changing the size
                return True
            except IndexError:
                message = "\n\n[vipdjwmsapp.davidjones.com.au - PuTTY] window is not active or not open."
                self.text_box.insert(tk.END, message)
                self.text_box.see(tk.END)  # Scroll to the bottom
                return False

        def is_notepad_plus_plus_active():
            try:
                notepad_window = gw.getWindowsWithTitle("vipdjwmsapp.davidjones.com.au - PuTTY")[0]
                return notepad_window.isActive
            except IndexError:
                return False  # Assume it's closed if not found

        if self.execution_running:
            message = "\n\nExecution is already in progress."
            self.text_box.insert(tk.END, message)
            self.text_box.see(tk.END)  # Scroll to the bottom
            return

        # 锁定下拉选择框
        self.sort_option_menu.config(state='disabled')
        self.date_options_menu.config(state='disabled')
        
        # 使用过滤后的DataFrame更新self.combined_df
        self.combined_df = self.filtered_df.copy()

        if not activate_notepad_plus_plus():
            return  # If vipdjwmsapp.davidjones.com.au - PuTTY cannot be activated, just return

        # Adding a message indicating that the execution has started
        start_message = "\n\nExecution has started. Processing locations..."
        self.text_box.insert(tk.END, start_message)
        self.text_box.see(tk.END)  # Ensure the message is visible

        self.export_button.config(text="Executing")
        self.execution_running = True

        pd.set_option('future.no_silent_downcasting', True)
        sleep_time = float(self.sleep_time_var.get()) / 1000.0
        locations = self.combined_df['Location'].apply(lambda x: str(x).zfill(9)).tolist()
        barcode_data_mark = False  # 在循环开始前定义
        base64_image_data = "iVBORw0KGgoAAAANSUhEUgAAADgAAAAICAYAAACs/DyzAAAAAXNSR0IArs4c6QAAAARnQU1BAACxjwv8YQUAAAAJcEhZcwAADsEAAA7BAbiRa+0AAAJpSURBVEhLnVWvjvswDP7uHqOKThrNA0zVSNFQWUHJQHHRxgaKC8Y2VFwwMlBWVDQyVXuA0JOmqK/R+9J/6+7W3353n2QpsWPHjh37LS6Keg6DCtk6wNlJcfCslpOtESQzxMUG7RmiyrAMkm5DuDGKzSBtUF0z7KIEims3LmDErS0FGd7t47rHMsqHMw/4JvuXPiARxlvYwoJFkbm/hA0cA8rcuE5DWXM1IlnHaVzLMU+GdZqGtRundShHfBIvHdmQPFM82DQ6Ke253b631ct7erRzp1f6Rh67Iz1p/G/9fMenBsSMfMlMFShimoHZ6yYDPaRjQ58S5BcN22Hok1DIj1ln846yBPzG9t8wqS9D+KJElI+8VQpREIAJx3vDEB8mAoiqwpVrKT8g9K0RtZBwbI2jqYb8CG075ExAsly2fIyLOTzCOcIJPviqf8OU/kww+nO3+Yl3qBtzRb+agyeubTgO1/qzPWHgrmDrS5dRhYu2sfr2mJZ3QGEq4LCFOO34op1ghPzIn7ENpx/nBV7ru20V9sSMM4MMxBJYLebQtxxnloJtM8ARXMqs+WZQ3MwtzBePEZomsFwusd5rCH/Fq55AJdiVfByn2/8Wz/SHL2aQI6IPxo/9lU2TJccAFXQlIMQVpqrUjQpsRfrW17SLhciw7hRbWiMTi6dBqDyiEwKbif+mkhNf0EPXB3+NH/oMuhQsXfcxr/x0Dd5IputhwyDa9m/S7ENzZCQIkR56Y+0YSZSR92ODvEzD87oeP4wQ/sP0AA8Z9tobRkDf6pvR4uth3PzPmDCY0m/u45jwWFkt2Ev25psofAGNAGgVBZyj1QAAAABJRU5ErkJggg=="
        image_data = base64.b64decode(base64_image_data)
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_image:
            tmp_image.write(image_data)
            tmp_image_path = tmp_image.name

        for index in range(self.current_location_index, len(locations)):
            print("Start Location：",self.current_location_index, "/", len(locations))
            if not is_notepad_plus_plus_active():
                message = f"\n\n[vipdjwmsapp.davidjones.com.au - PuTTY] window is not active or not open.\nExecution paused at Location: {locations[self.current_location_index-1]}.\nPlease return to Cycle Count entry, then continue execution."
                self.text_box.insert(tk.END, message)
                self.text_box.see(tk.END)  # Scroll to the bottom
                break

            barcode_data_mark = False
            current_location = locations[self.current_location_index]
            location_df = self.combined_df[self.combined_df['Location'] == current_location]
            print(location_df)
            
            if location_df['Current Qty'].str.match(r'^\d+\s+units$').any(): # 'Current Qty'的值符合“数字+空格+units”的格式
                barcode_data_mark = True # 说明有barcodes需要处理
                location_df.loc[:, 'Barcode'] = location_df['Barcode'].replace('nan', np.nan)# 确保 'Barcode' 列中的 'nan' 文本被正确视为 NaN
                self.current_location_index += len(location_df)
                
                if location_df['Barcode'].isna().any() or location_df['Expiry Date'].notna().any(): # 检查 'Barcode' 为空或 'Expiry Date' 不为空的情况
                    message = f"\n\n{locations[self.current_location_index-1]} has been skipped, due to empty barcode / expiry date."
                    self.text_box.insert(tk.END, message)
                    self.text_box.see(tk.END)  # Scroll to the bottom
                    continue  # 可以直接去下一个循环

            else:
                self.current_location_index += 1 # 如果当前location没有Current Qty，即emplty location

            pyautogui.write(current_location)
            # print ("Location", current_location)
            pyautogui.press('enter')
            time.sleep(sleep_time)
            try:
                image_location = pyautogui.locateCenterOnScreen(tmp_image_path, confidence=0.5)
                if image_location:
                    pyautogui.hotkey('ctrl', 'a')
                    time.sleep(sleep_time)
                    # print ("Ctrl+A, 等待", sleep_time, "秒")
            except Exception as e:
                print("发生错误：", e)

            # 准备输入barcodes
            if barcode_data_mark:
                current_location_df = self.combined_df[self.combined_df['Location'] == current_location] # 筛选当前Location的数据
                for _, row in current_location_df.iterrows(): # 遍历current_location_df这个DataFrame中的每一行
                    current_qty_data = row['Current Qty'] # 从Current Qty列中提取数量
                    current_bc_qty = int(current_qty_data.split(' ')[0])  # 从格式是数字+空格+units中提取数量
                    current_bc = str(row['Barcode']).zfill(13)  # 从Barcode列中提取条形码并格式化, 在前面加0到达13位

                    # 对于每个条形码，执行指定次数的发送操作
                    for _ in range(current_bc_qty):
                        pyautogui.write(current_bc)
                        # print ("Barcode", current_bc)
                        pyautogui.press('enter')
                        time.sleep(sleep_time)

            pyautogui.hotkey('ctrl', 'n') # 实际操作时激活这一行
            # pyautogui.hotkey('ctrl', 'x') # 测试时激活这一行, 输入完BC后直接退出 
            # time.sleep(sleep_time) # 测试时激活这一行, 输入完BC后直接退出
            # pyautogui.write('7') # 测试时激活这一行, 输入完BC后直接退出
            # pyautogui.press('enter') # 测试时激活这一行, 输入完BC后直接退出
            time.sleep(sleep_time)
            # print ("Ctrl+N, 等待", sleep_time, "秒")

            self.update()

            if self.current_location_index==len(locations):  # If the loop finishes normally without a break
                break

        self.current_location_index = 0  # Reset position for next execution
        complete_message = "\n\nExecution complete. All locations have been processed."
        self.text_box.insert(tk.END, complete_message)
        self.text_box.see(tk.END)
        self.export_button.config(text="Start Execution")
        self.execution_running = False

if __name__ == "__main__":
    app = MainWindow()
    app.mainloop()
